import datetime
import random

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

class Demo001:

    @staticmethod
    def demo1():
        print('----------------读取Excel文件-------------------')
        # 加载一个工作簿 ---> Workbook
        wb = openpyxl.load_workbook('temp_file/二审确定名单(1).xlsx')
        # 获取工作表的名字
        print(wb.sheetnames)
        # 获取工作表 ---> Worksheet
        sheet = wb.worksheets[0]
        # 获得单元格的范围
        print(sheet.dimensions)
        # 获得行数和列数
        print(sheet.max_row, sheet.max_column)

        # 获取指定单元格的值
        print(sheet.cell(3, 3).value)
        print(sheet['C3'].value)
        print(sheet['G255'].value)

        # 获取多个单元格（嵌套元组）
        print(sheet['A2:C5'])

        print('读取文件所有内容：')
        for row_ch in range(2, sheet.max_row + 1):
            for col_ch in 'ABCDEFG':
                value = sheet[f'{col_ch}{row_ch}'].value
                print(value, end='\t')
            print()

    @staticmethod
    def demo2():
        print('----------------写Excel文件-------------------')
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '期末成绩'
        titles = ('姓名', '语文', '数学', '英语')
        for col_index, title in enumerate(titles):
            sheet.cell(1, col_index + 1, title)

        names = ('关羽', '张飞', '赵云', '马超', '黄忠')
        for row_index, name in enumerate(names):
            sheet.cell(row_index + 2, 1, name)
            for col_index in range(2, 5):
                sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

        wb.save('temp_file/考试成绩单2.xlsx')
        print('文件写入成功')
    @staticmethod
    def demo3():
        print('----------------写Excel文件|整单元格样式-------------------')
        # 对齐方式
        alignment = Alignment(horizontal='center', vertical='center')
        # 边框线条
        side = Side(color='ff7f50', style='mediumDashed')

        wb = openpyxl.load_workbook('temp_file/考试成绩单2.xlsx')
        sheet = wb.worksheets[0]

        # 调整行高和列宽
        sheet.row_dimensions[1].height = 30
        sheet.column_dimensions['E'].width = 120

        sheet['E1'] = '平均分'
        # 设置字体
        sheet.cell(1, 5).font = Font(size=18, bold=True, color='ff1493', name='华文楷体')
        # 设置对齐方式
        sheet.cell(1, 5).alignment = alignment
        # 设置单元格边框
        sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
        for i in range(2, 7):
            # 公式计算每个学生的平均分
            sheet[f'E{i}'] = f'=average(B{i}:D{i})'
            sheet.cell(i, 5).font = Font(size=12, color='4169e1', italic=True)
            sheet.cell(i, 5).alignment = alignment


        wb.save('temp_file/考试成绩单2.xlsx')
        print('文件写入成功')
    @staticmethod
    def demo4():
        print('----------------生成统计图表-------------------')

        wb = Workbook(write_only=True)
        sheet = wb.create_sheet()

        rows = [
            ('类别', '销售A组', '销售B组'),
            ('手机', 40, 30),
            ('平板', 50, 60),
            ('笔记本', 80, 70),
            ('外围设备', 20, 10),
        ]

        # 向表单中添加行
        for row in rows:
            sheet.append(row)

        # 创建图表对象
        chart = BarChart()
        chart.type = 'col'
        chart.style = 10
        # 设置图表的标题
        chart.title = '销售统计图'
        # 设置图表纵轴的标题
        chart.y_axis.title = '销量'
        # 设置图表横轴的标题
        chart.x_axis.title = '商品类别'
        # 设置数据的范围
        data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)
        # 设置分类的范围
        cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
        # 给图表添加数据
        chart.add_data(data, titles_from_data=True)
        # 给图表设置分类
        chart.set_categories(cats)
        chart.shape = 4
        # 将图表添加到表单指定的单元格中
        sheet.add_chart(chart, 'A10')

        wb.save('temp_file/demo.xlsx')


    @staticmethod
    def demo5():
        print('----------------demo-------------------')

    @staticmethod
    def demo6():
        print('----------------demo-------------------')


if __name__ == '__main__':
    # Demo001.demo1()
    # Demo001.demo2()
    # Demo001.demo3()
    # 生成统计图表
    Demo001.demo4()
